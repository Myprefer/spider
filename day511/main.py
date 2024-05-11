import random
from time import sleep

import openpyxl
import requests

filename = "company_info.xlsx"

try:
    wb = openpyxl.load_workbook(filename)
except FileNotFoundError:
    wb = openpyxl.Workbook()

ws = wb.active

if ws.max_row == 1:
    ws.append([
        "Company ID",
        "Company Name (English)",
        "Company Name (Legal)",
        "Brand",
        "Identifier",
        "Suspicious",
        "Latitude",
        "Longitude",
        "Full Address",
        "Transactions Involved",
        "Source",
        "Geocode Level",
        "Country",
        "Investigation Status",
        "Locations",
        "Address",
        "Top Suppliers",
        "Top Customers",
        "Top Acquired Products",
        "Top Supplied Products",
        "Year Data",
        "Top Supplying Countries",
        "Top Buying Countries",
        "Evidences"
    ])


def get_all_company_info(id):
    response = requests.get(f'https://api.supplytrace.org/api/getAllCompanyInfo/?search={id}', headers=headers)
    if response.status_code == 200:
        json_form = response.json()
        # print(json_form)

        all_company_info = {
            'address': json_form.get('address'),
            'top_suppliers': json_form.get('top_suppliers'),
            'top_customers': json_form.get('top_customers'),
            'top_acquired_products': json_form.get('top_acquired_products'),
            'top_supplied_products': json_form.get('top_supplied_products'),
            'year_data': json_form.get('year_data'),
            'top_supplying_countries': json_form.get('top_supplying_countries'),
            'top_buying_countries': json_form.get('top_buying_countries'),
        }
        # print(all_company_info)
        return all_company_info


def company_locations(id):
    response = requests.get(f'https://api.supplytrace.org/api/loadCompanyLocations/?company={id}', headers=headers)
    if response.status_code == 200:
        main_info = response.json().get('company')
        # print(main_info)

        basic_info = {
            "id": main_info.get('id'),
            "company_name_english": main_info.get('company_name_english'),
            "country": main_info.get('country'),
            "address": main_info.get('full_address'),
            "brand": main_info.get('brand'),
            "company_name_legal": main_info.get('company_name_legal'),
            "source": main_info.get('source'),
            "identifier": main_info.get('identifier'),
            "suspicious": main_info.get('suspicious'),
            "latitude": main_info.get('latitude'),
            "longitude": main_info.get('longitude'),
            "transactions_involved": main_info.get('transactions_involved'),
            "geocode_level": main_info.get('geocode_level'),
            "investigation_status": main_info.get('investigation_status')
        }

        locations = response.json().get('locations')
        for location in locations:
            location = {
                'id': location.get('id'),
                'company_name_english': location.get('company_name_english'),
                'suspicious': location.get('suspicious'),
                'full_address': location.get('full_address'),
                'latitude': location.get('latitude'),
                'longitude': location.get('longitude')
            }
    # print(basic_info)
    # print(locations)
    return basic_info, locations


headers = {
    'accept': 'application/json, text/plain, */*',
    'accept-language': 'zh-CN,zh;q=0.9',
    'authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ0b2tlbl90eXBlIjoiYWNjZXNzIiwiZXhwIjoxNzE1NTAwNDA4LCJpYXQiOjE3MTU0MTQwMDgsImp0aSI6IjhiZTdlZmI4MGQwMjQ0YTI4ZmY2NzYwOWQ2NmQxYWNmIiwidXNlcl9pZCI6NzgyfQ.Q78I9vrSbLzZ46OFo_Smd7ibqpYW2Shy-Dn3uSZZ1_s',
    'cache-control': 'no-cache',
    'origin': 'https://supplytrace.org',
    'pragma': 'no-cache',
    'priority': 'u=1, i',
    'referer': 'https://supplytrace.org/',
    'sec-ch-ua': '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-site',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
}


def process_line(line):
    id = int(line.strip())
    print(id)

    company, locations = company_locations(id)
    all_company_info = get_all_company_info(id)

    ws.append([
        str(company.get("id")),
        str(company.get("company_name_english")),
        str(company.get("company_name_legal")),
        str(company.get("brand")),
        str(company.get("identifier")),
        str(company.get("suspicious")),
        str(company.get("latitude")),
        str(company.get("longitude")),
        str(company.get("full_address")),
        str(company.get("transactions_involved")),
        str(company.get("source")),
        str(company.get("geocode_level")),
        str(company.get("country")),
        str(company.get("investigation_status")),
        str(locations),
        str(all_company_info.get('address')),
        str(all_company_info.get('top_suppliers')),
        str(all_company_info.get('top_customers')),
        str(all_company_info.get('top_acquired_products')),
        str(all_company_info.get('top_supplied_products')),
        str(all_company_info.get('year_data')),
        str(all_company_info.get('top_supplying_countries')),
        str(all_company_info.get('top_buying_countries')),
        str(all_company_info.get('evidences'))
    ])

    wb.save("company_info.xlsx")
    sleep(random.uniform(0.5, 1.0))


for char in 'abcdefghijklmnopqrstuvwxyz':
    with open(f'./id/{char}.txt', 'r', encoding='utf-8') as f:
        lines = f.readlines()
        for line in lines:
            process_line(line)
